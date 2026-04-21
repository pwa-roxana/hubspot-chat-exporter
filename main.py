from flask import Flask, request, jsonify, send_file, render_template_string, session, redirect, Response
import requests as req
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io, os, json

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "pwa-chat-exporter-secret")
HUBSPOT_TOKEN = os.environ.get("HUBSPOT_TOKEN", "")
APP_PASSWORD = os.environ.get("APP_PASSWORD", "pwa2024")
BASE = "https://api.hubapi.com"
agent_cache = {}

KNOWN_AGENTS = {
    'A-10327929': 'Maricruz Alonso',
    'A-79088480': 'Colin Moser',
    'A-12160092': 'John Gill',
}

LOGIN_HTML = """<!DOCTYPE html><html><head><title>PWA Live Chat</title>
<meta name="viewport" content="width=device-width,initial-scale=1">
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f4f6f9;min-height:100vh;display:flex;align-items:center;justify-content:center}
.card{background:#fff;border-radius:12px;padding:2rem;width:100%;max-width:380px;box-shadow:0 2px 16px rgba(0,0,0,.08)}
.logo{font-size:11px;font-weight:700;color:#f97316;letter-spacing:.1em;text-transform:uppercase;margin-bottom:.5rem}
h1{font-size:20px;font-weight:600;color:#1a1a2e;margin-bottom:.25rem}
p{font-size:13px;color:#6b7280;margin-bottom:1.5rem}
label{display:block;font-size:13px;font-weight:500;color:#374151;margin-bottom:5px}
input{width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;margin-bottom:1rem;outline:none}
input:focus{border-color:#1F3864}
button{width:100%;padding:11px;background:#1F3864;color:#fff;border:none;border-radius:8px;font-size:15px;font-weight:600;cursor:pointer}
.err{background:#fef2f2;color:#991b1b;border:1px solid #fecaca;border-radius:8px;padding:10px 12px;font-size:13px;margin-bottom:1rem}
</style></head><body>
<div class="card">
  <div class="logo">Pacific West Academy</div>
  <h1>Live Chat Intelligence</h1>
  <p>Enter the team password to access the dashboard</p>
  {% if error %}<div class="err">{{ error }}</div>{% endif %}
  <form method="POST" action="/login">
    <label>Password</label>
    <input type="password" name="password" placeholder="Enter password" autofocus/>
    <button type="submit">Sign in</button>
  </form>
</div></body></html>"""

DASHBOARD_HTML = """<!DOCTYPE html><html><head><title>PWA Live Chat Dashboard</title>
<meta name="viewport" content="width=device-width,initial-scale=1">
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f0f2f5;color:#1a1a2e}
.topbar{background:#1F3864;padding:14px 24px;display:flex;align-items:center;justify-content:space-between}
.topbar-left{display:flex;align-items:center;gap:16px}
.logo{font-size:11px;font-weight:700;color:#f97316;letter-spacing:.1em;text-transform:uppercase}
.topbar h1{font-size:16px;font-weight:600;color:#fff;margin:0}
.topbar-right{display:flex;align-items:center;gap:12px}
.topbar a{font-size:12px;color:rgba(255,255,255,.6);text-decoration:none}
.topbar a:hover{color:#fff}
.filters{background:#fff;border-bottom:1px solid #e5e7eb;padding:12px 24px;display:flex;align-items:center;gap:12px;flex-wrap:wrap}
.filters label{font-size:12px;color:#6b7280;font-weight:500}
.filters input{padding:6px 10px;border:1px solid #d1d5db;border-radius:6px;font-size:13px;outline:none}
.filters input:focus{border-color:#1F3864}
.load-btn{padding:8px 20px;background:#1F3864;color:#fff;border:none;border-radius:6px;font-size:13px;font-weight:600;cursor:pointer}
.load-btn:disabled{background:#9ca3af;cursor:not-allowed}
.export-btn{padding:8px 16px;background:#fff;color:#1F3864;border:1px solid #1F3864;border-radius:6px;font-size:13px;font-weight:600;cursor:pointer}
.main{padding:20px 24px;max-width:1200px;margin:0 auto}
.metrics{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:12px;margin-bottom:20px}
.metric{background:#fff;border-radius:10px;padding:16px;border:0.5px solid #e5e7eb}
.metric-label{font-size:11px;color:#6b7280;text-transform:uppercase;letter-spacing:.05em;margin-bottom:6px}
.metric-value{font-size:26px;font-weight:600;color:#1a1a2e}
.metric-sub{font-size:11px;color:#9ca3af;margin-top:3px}
.grid2{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px}
.grid3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:16px;margin-bottom:16px}
.card{background:#fff;border-radius:10px;padding:16px;border:0.5px solid #e5e7eb}
.card-title{font-size:11px;font-weight:600;color:#6b7280;text-transform:uppercase;letter-spacing:.06em;margin-bottom:14px}
.chart-wrap{position:relative}
.bar-row{display:flex;align-items:center;gap:8px;margin-bottom:8px}
.bar-label{font-size:12px;color:#374151;min-width:150px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.bar-track{flex:1;height:8px;background:#f0f2f5;border-radius:4px;overflow:hidden}
.bar-fill{height:100%;border-radius:4px}
.bar-count{font-size:12px;color:#6b7280;min-width:28px;text-align:right}
table{width:100%;border-collapse:collapse;font-size:13px}
th{text-align:left;font-size:11px;font-weight:600;color:#6b7280;padding:8px;border-bottom:1px solid #e5e7eb;text-transform:uppercase;letter-spacing:.04em}
td{padding:10px 8px;border-bottom:0.5px solid #f0f2f5;color:#374151;vertical-align:top}
tr:last-child td{border-bottom:none}
tr:hover td{background:#f9fafb}
.badge{display:inline-block;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:600}
.badge-closed{background:#dcfce7;color:#166534}
.badge-open{background:#fef3c7;color:#92400e}
.status-bar{background:#1F3864;color:#fff;padding:10px 24px;font-size:13px;display:none;align-items:center;gap:10px}
.spinner{width:14px;height:14px;border:2px solid rgba(255,255,255,.3);border-top-color:#fff;border-radius:50%;animation:spin .7s linear infinite;flex-shrink:0}
@keyframes spin{to{transform:rotate(360deg)}}
.ai-box{background:#f0f4ff;border:1px solid #c7d2fe;border-radius:10px;padding:16px;margin-bottom:16px}
.ai-title{font-size:11px;font-weight:600;color:#3730a3;text-transform:uppercase;letter-spacing:.06em;margin-bottom:10px}
.ai-item{display:flex;gap:8px;margin-bottom:8px;align-items:flex-start}
.ai-dot{width:5px;height:5px;border-radius:50%;background:#4f46e5;margin-top:6px;flex-shrink:0}
.ai-text{font-size:13px;color:#1e1b4b;line-height:1.5}
.ai-btn{padding:8px 16px;background:#4f46e5;color:#fff;border:none;border-radius:6px;font-size:13px;font-weight:600;cursor:pointer;margin-top:8px}
.ai-btn:disabled{opacity:.5;cursor:not-allowed}
.empty{text-align:center;padding:2rem;color:#9ca3af;font-size:14px}
.legend{display:flex;flex-wrap:wrap;gap:12px;margin-bottom:8px;font-size:11px;color:#6b7280}
.legend span{display:flex;align-items:center;gap:4px}
.leg-dot{width:10px;height:10px;border-radius:2px;flex-shrink:0}
@media(max-width:640px){.grid2,.grid3{grid-template-columns:1fr}}
</style></head><body>

<div class="topbar">
  <div class="topbar-left">
    <div>
      <div class="logo">Pacific West Academy</div>
      <h1>Live Chat Intelligence Dashboard</h1>
    </div>
  </div>
  <div class="topbar-right">
    <span id="last-refresh" style="font-size:11px;color:rgba(255,255,255,.5)"></span>
    <a href="/logout">Sign out</a>
  </div>
</div>

<div class="filters">
  <label>From</label>
  <input type="date" id="from-date"/>
  <label>To</label>
  <input type="date" id="to-date"/>
  <button class="load-btn" id="load-btn" onclick="loadData()">Load data</button>
  <button class="export-btn" onclick="exportExcel()">Export Excel</button>
</div>

<div class="status-bar" id="status-bar">
  <div class="spinner"></div>
  <span id="status-msg">Loading conversations...</span>
</div>

<div class="main" id="main-content" style="display:none">
  <div class="metrics" id="metrics-row"></div>

  <div class="grid3">
    <div class="card">
      <div class="card-title">Conversations by day</div>
      <div class="chart-wrap" style="height:180px"><canvas id="dow-chart" role="img" aria-label="Conversations by day of week"></canvas></div>
    </div>
    <div class="card">
      <div class="card-title">Working vs. off hours</div>
      <div class="legend" id="hours-legend"></div>
      <div class="chart-wrap" style="height:155px"><canvas id="hours-chart" role="img" aria-label="Working vs off hours"></canvas></div>
    </div>
    <div class="card">
      <div class="card-title">Status breakdown</div>
      <div class="legend" id="status-legend"></div>
      <div class="chart-wrap" style="height:155px"><canvas id="status-chart" role="img" aria-label="Status breakdown"></canvas></div>
    </div>
  </div>

  <div class="card" style="margin-bottom:16px">
    <div class="card-title">Hourly activity (working hours highlighted)</div>
    <div class="chart-wrap" style="height:180px"><canvas id="hourly-chart" role="img" aria-label="Hourly activity"></canvas></div>
  </div>

  <div class="card" style="margin-bottom:16px">
    <div class="card-title">Conversations over time</div>
    <div class="chart-wrap" style="height:180px"><canvas id="timeline-chart" role="img" aria-label="Conversations over time"></canvas></div>
  </div>

  <div class="grid2">
    <div class="card">
      <div class="card-title">Top topics & inquiries</div>
      <div id="topics-list"></div>
    </div>
    <div class="card">
      <div class="card-title">Agent activity</div>
      <div id="agents-list"></div>
    </div>
  </div>

  <div class="ai-box" id="ai-box">
    <div class="ai-title">AI executive analysis</div>
    <div id="ai-content"><p style="font-size:13px;color:#6b7280">Click below to run AI analysis on your conversations.</p></div>
    <button class="ai-btn" id="ai-btn" onclick="runAI()">Analyze with AI</button>
  </div>

  <div class="card">
    <div class="card-title">Recent conversations</div>
    <div style="overflow-x:auto">
      <table>
        <thead><tr>
          <th style="width:90px">Date</th>
          <th style="width:140px">Visitor</th>
          <th style="width:120px">Assigned to</th>
          <th style="width:80px">Status</th>
          <th>Last message</th>
        </tr></thead>
        <tbody id="conv-tbody"></tbody>
      </table>
    </div>
  </div>
</div>

<script>
const WORK_START=8,WORK_END=18;
const DAYS=['Sun','Mon','Tue','Wed','Thu','Fri','Sat'];
let charts={},allRows=[],parsedConvos=[];

function destroyChart(id){if(charts[id]){charts[id].destroy();delete charts[id];}}

function setStatus(msg,show=true){
  const bar=document.getElementById('status-bar');
  document.getElementById('status-msg').textContent=msg;
  bar.style.display=show?'flex':'none';
}

async function loadData(){
  const btn=document.getElementById('load-btn');
  btn.disabled=true;
  setStatus('Fetching conversations from HubSpot...');
  document.getElementById('main-content').style.display='none';
  const from=document.getElementById('from-date').value||'';
  const to=document.getElementById('to-date').value||'';
  try{
    const params=new URLSearchParams();
    if(from)params.set('from_date',from);
    if(to)params.set('to_date',to);
    const resp=await fetch('/api/data?'+params.toString());
    if(!resp.ok){const e=await resp.json();throw new Error(e.error||'Failed');}
    allRows=await resp.json();
    setStatus('Processing data...');
    processData(allRows);
    document.getElementById('main-content').style.display='block';
    document.getElementById('last-refresh').textContent='Last updated '+new Date().toLocaleTimeString();
    setStatus('',false);
  }catch(e){
    setStatus('Error: '+e.message);
    setTimeout(()=>setStatus('',false),4000);
  }finally{btn.disabled=false;}
}

async function exportExcel(){
  const from=document.getElementById('from-date').value||null;
  const to=document.getElementById('to-date').value||null;
  setStatus('Generating Excel file...');
  const resp=await fetch('/export',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({from_date:from,to_date:to})});
  if(resp.ok){
    const blob=await resp.blob();
    const url=URL.createObjectURL(blob);
    const a=document.createElement('a');
    a.href=url;a.download='hubspot_live_chats_'+new Date().toISOString().slice(0,10)+'.xlsx';a.click();
  }
  setStatus('',false);
}

function parseDate(s){if(!s)return null;const d=new Date(s);return isNaN(d)?null:d;}

function processData(rows){
  const threadMap={};
  rows.forEach(r=>{
    const tid=r.thread_id||'';
    if(!threadMap[tid])threadMap[tid]={
      id:tid,date:parseDate(r.created_at),status:r.status||'',
      visitor:r.visitor_name||'',email:r.visitor_email||'',
      assigned:r.assigned_to||'',messages:[]
    };
    if(r.message)threadMap[tid].messages.push({time:parseDate(r.message_time),sender:r.sender||'',text:r.message});
  });
  parsedConvos=Object.values(threadMap).filter(c=>c.date).sort((a,b)=>b.date-a.date);
  renderMetrics(parsedConvos);
  renderDOW(parsedConvos);
  renderHours(parsedConvos);
  renderStatus(parsedConvos);
  renderHourly(parsedConvos);
  renderTimeline(parsedConvos);
  renderTopics(parsedConvos);
  renderAgents(parsedConvos);
  renderTable(parsedConvos);
}

function renderMetrics(c){
  const total=c.length;
  const closed=c.filter(x=>(x.status||'').toLowerCase().includes('closed')).length;
  const withEmail=c.filter(x=>x.email).length;
  const offHours=c.filter(x=>{const h=x.date.getHours();return h<WORK_START||h>=WORK_END;}).length;
  const totalMsgs=c.reduce((s,x)=>s+x.messages.length,0);
  const avgMsgs=total?Math.round(totalMsgs/total):0;
  const metrics=[
    {label:'Total conversations',value:total,sub:'all time'},
    {label:'Resolved',value:closed,sub:total?Math.round(closed/total*100)+'% resolution rate':''},
    {label:'Contact captured',value:withEmail,sub:total?Math.round(withEmail/total*100)+'% of visitors':''},
    {label:'Off-hours chats',value:offHours,sub:total?Math.round(offHours/total*100)+'% of total':''},
    {label:'Avg messages',value:avgMsgs,sub:'per conversation'},
  ];
  document.getElementById('metrics-row').innerHTML=metrics.map(m=>`
    <div class="metric">
      <div class="metric-label">${m.label}</div>
      <div class="metric-value">${m.value}</div>
      <div class="metric-sub">${m.sub}</div>
    </div>`).join('');
}

function renderDOW(c){
  destroyChart('dow-chart');
  const counts=Array(7).fill(0);
  c.forEach(x=>counts[x.date.getDay()]++);
  charts['dow-chart']=new Chart(document.getElementById('dow-chart'),{
    type:'bar',
    data:{labels:DAYS,datasets:[{data:counts,backgroundColor:counts.map((_,i)=>i===0||i===6?'#93c5fd':'#1F3864'),borderRadius:4}]},
    options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},
      scales:{x:{grid:{display:false},ticks:{font:{size:11}}},y:{grid:{color:'#f0f2f5'},ticks:{font:{size:11},stepSize:1}}}}
  });
}

function renderHours(c){
  destroyChart('hours-chart');
  let work=0,off=0;
  c.forEach(x=>{const h=x.date.getHours();if(h>=WORK_START&&h<WORK_END)work++;else off++;});
  document.getElementById('hours-legend').innerHTML=`
    <span><span class="leg-dot" style="background:#1F3864"></span>Work hrs ${work}</span>
    <span><span class="leg-dot" style="background:#93c5fd"></span>Off hrs ${off}</span>`;
  charts['hours-chart']=new Chart(document.getElementById('hours-chart'),{
    type:'doughnut',
    data:{labels:['Working hours','Off hours'],datasets:[{data:[work,off],backgroundColor:['#1F3864','#93c5fd'],borderWidth:0}]},
    options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},cutout:'65%'}
  });
}

function renderStatus(c){
  destroyChart('status-chart');
  const counts={};
  c.forEach(x=>{const s=(x.status||'Unknown').replace(/_/g,' ');counts[s]=(counts[s]||0)+1;});
  const labels=Object.keys(counts),data=Object.values(counts);
  const colors=['#1F3864','#93c5fd','#6b7280','#fbbf24'];
  document.getElementById('status-legend').innerHTML=labels.map((l,i)=>`
    <span><span class="leg-dot" style="background:${colors[i%colors.length]}"></span>${l} ${data[i]}</span>`).join('');
  charts['status-chart']=new Chart(document.getElementById('status-chart'),{
    type:'doughnut',
    data:{labels,datasets:[{data,backgroundColor:colors,borderWidth:0}]},
    options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},cutout:'60%'}
  });
}

function renderHourly(c){
  destroyChart('hourly-chart');
  const counts=Array(24).fill(0);
  c.forEach(x=>counts[x.date.getHours()]++);
  const labels=Array.from({length:24},(_,i)=>i===0?'12am':i<12?i+'am':i===12?'12pm':(i-12)+'pm');
  charts['hourly-chart']=new Chart(document.getElementById('hourly-chart'),{
    type:'bar',
    data:{labels,datasets:[{data:counts,backgroundColor:counts.map((_,i)=>i>=WORK_START&&i<WORK_END?'#1F3864':'#93c5fd'),borderRadius:3}]},
    options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},
      scales:{x:{grid:{display:false},ticks:{font:{size:10},maxRotation:45,autoSkip:false}},
        y:{grid:{color:'#f0f2f5'},ticks:{font:{size:11},stepSize:1}}}}
  });
}

function renderTimeline(c){
  destroyChart('timeline-chart');
  const byDate={};
  c.forEach(x=>{const k=x.date.toISOString().slice(0,10);byDate[k]=(byDate[k]||0)+1;});
  const sorted=Object.keys(byDate).sort();
  charts['timeline-chart']=new Chart(document.getElementById('timeline-chart'),{
    type:'line',
    data:{labels:sorted,datasets:[{label:'Conversations',data:sorted.map(k=>byDate[k]),
      borderColor:'#1F3864',backgroundColor:'rgba(31,56,100,0.08)',fill:true,tension:.3,pointRadius:3,pointBackgroundColor:'#1F3864'}]},
    options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},
      scales:{x:{grid:{display:false},ticks:{font:{size:11},maxTicksLimit:10}},
        y:{grid:{color:'#f0f2f5'},ticks:{font:{size:11},stepSize:1}}}}
  });
}

function renderTopics(c){
  const kws={
    'Eligibility / requirements':['eligib','require','qualify','felon','background','age','citizen'],
    'Program info / courses':['program','course','class','train','certif','cess','cst','curriculum'],
    'Location / campus':['locat','campus','where','address','van nuys','san diego','southern ca'],
    'Scheduling / start dates':['start','when','schedul','next','upcoming','enroll','begin'],
    'Pricing / cost / GI Bill':['cost','price','pay','tuition','financ','gi bill','afford','how much'],
    'Job placement / careers':['job','career','hire','placement','employ','salary'],
    'CCTV / armed / guard card':['cctv','camera','firearm','armed','unarmed','guard card','surveil'],
    'Contact / follow up':['contact','call','email','reach','phone','speak','follow'],
  };
  const counts={};Object.keys(kws).forEach(k=>counts[k]=0);
  c.forEach(x=>{
    const txt=x.messages.map(m=>m.text).join(' ').toLowerCase();
    Object.entries(kws).forEach(([topic,words])=>{if(words.some(w=>txt.includes(w)))counts[topic]++;});
  });
  const sorted=Object.entries(counts).sort((a,b)=>b[1]-a[1]).filter(([,v])=>v>0);
  const max=sorted[0]?.[1]||1;
  document.getElementById('topics-list').innerHTML=sorted.map(([label,count])=>`
    <div class="bar-row">
      <div class="bar-label">${label}</div>
      <div class="bar-track"><div class="bar-fill" style="width:${Math.round(count/max*100)}%;background:#1F3864"></div></div>
      <div class="bar-count">${count}</div>
    </div>`).join('')||'<div class="empty">No topic data</div>';
}

function renderAgents(c){
  const counts={};
  c.forEach(x=>{const a=(x.assigned||'').trim()||'Unassigned';counts[a]=(counts[a]||0)+1;});
  const sorted=Object.entries(counts).sort((a,b)=>b[1]-a[1]);
  const max=sorted[0]?.[1]||1;
  document.getElementById('agents-list').innerHTML=sorted.map(([name,count])=>`
    <div class="bar-row">
      <div class="bar-label">${name.length>24?name.slice(0,24)+'...':name}</div>
      <div class="bar-track"><div class="bar-fill" style="width:${Math.round(count/max*100)}%;background:#059669"></div></div>
      <div class="bar-count">${count}</div>
    </div>`).join('');
}

function renderTable(c){
  document.getElementById('conv-tbody').innerHTML=c.slice(0,50).map((x,i)=>{
    const last=x.messages.filter(m=>m.text).slice(-1)[0];
    const isClosed=(x.status||'').toLowerCase().includes('closed');
    const badge=isClosed?
      `<span class="badge badge-closed" style="cursor:pointer" onclick="openModal(${i})">Closed</span>`:
      `<span class="badge badge-open" style="cursor:pointer" onclick="openModal(${i})">Open</span>`;
    return`<tr style="cursor:pointer" onclick="openModal(${i})">
      <td>${x.date.toLocaleDateString()}</td>
      <td title="${x.email}">${x.visitor||x.email||'Unknown'}</td>
      <td>${(x.assigned||'—').slice(0,20)}</td>
      <td>${badge}</td>
      <td style="font-size:12px;color:#6b7280;white-space:normal">${last?last.text.slice(0,100)+(last.text.length>100?'...':''):'—'}</td>
    </tr>`;
  }).join('');
}

function openModal(idx){
  const c=parsedConvos[idx];
  if(!c)return;
  const isClosed=(c.status||'').toLowerCase().includes('closed');
  const statusBadge=isClosed?
    '<span class="badge badge-closed">Closed</span>':
    '<span class="badge badge-open">Open</span>';
  const msgs=c.messages.filter(m=>m.text).map(m=>{
    const isAgent=m.sender&&m.sender!=='MESSAGE'&&!m.sender.includes('@');
    const align=isAgent?'right':'left';
    const bg=isAgent?'#1F3864':'#f0f2f5';
    const color=isAgent?'#fff':'#1a1a2e';
    const name=isAgent?(m.sender||'Agent'):'Visitor';
    const time=m.time?m.time.toLocaleTimeString([],{hour:'2-digit',minute:'2-digit'}):'';
    return`<div style="display:flex;flex-direction:column;align-items:${isAgent?'flex-end':'flex-start'};margin-bottom:12px">
      <div style="font-size:11px;color:#9ca3af;margin-bottom:3px">${name} ${time}</div>
      <div style="max-width:75%;padding:10px 14px;border-radius:12px;background:${bg};color:${color};font-size:13px;line-height:1.5">${m.text}</div>
    </div>`;
  }).join('');
  document.getElementById('modal-visitor').textContent=c.visitor||c.email||'Unknown visitor';
  document.getElementById('modal-email').textContent=c.email||'';
  document.getElementById('modal-date').textContent=c.date.toLocaleDateString('en-US',{weekday:'short',year:'numeric',month:'short',day:'numeric'});
  document.getElementById('modal-status').innerHTML=statusBadge;
  document.getElementById('modal-agent').textContent=c.assigned||'—';
  document.getElementById('modal-msgs').innerHTML=msgs||'<p style="color:#9ca3af;font-size:13px">No messages</p>';
  document.getElementById('conv-modal').style.display='flex';
}

function closeModal(){document.getElementById('conv-modal').style.display='none';}

async function runAI(){
  const btn=document.getElementById('ai-btn');
  btn.disabled=true;btn.textContent='Analyzing...';
  const content=document.getElementById('ai-content');
  content.innerHTML='<p style="font-size:13px;color:#6b7280">Running AI analysis...</p>';
  const sample=parsedConvos.slice(0,25).map(c=>({
    date:c.date.toISOString().slice(0,10),visitor:c.visitor||'Unknown',status:c.status,
    messages:c.messages.slice(0,6).map(m=>m.sender+': '+m.text).join(' | ')
  }));
  try{
    const resp=await fetch('/ai-analysis',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({conversations:sample})});
    const data=await resp.json();
    const lines=(data.analysis||'').split('\\n').filter(l=>l.trim());
    content.innerHTML=lines.map(line=>{
      const clean=line.replace(/^[\\d\\.\\-\\*]+\\s*/,'').replace(/\\*\\*/g,'').trim();
      if(!clean)return'';
      return`<div class="ai-item"><div class="ai-dot"></div><div class="ai-text">${clean}</div></div>`;
    }).join('');
  }catch(e){content.innerHTML='<p style="font-size:13px;color:#991b1b">Analysis failed: '+e.message+'</p>';}
  finally{btn.disabled=false;btn.textContent='Re-analyze';}
}
</script>

<div id="conv-modal" style="display:none;position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,0.5);z-index:1000;align-items:center;justify-content:center;padding:20px">
  <div style="background:#fff;border-radius:12px;width:100%;max-width:600px;max-height:85vh;display:flex;flex-direction:column">
    <div style="padding:16px 20px;border-bottom:1px solid #e5e7eb;display:flex;align-items:flex-start;justify-content:space-between">
      <div>
        <div style="font-size:16px;font-weight:600;color:#1a1a2e" id="modal-visitor"></div>
        <div style="font-size:12px;color:#6b7280;margin-top:2px" id="modal-email"></div>
        <div style="display:flex;align-items:center;gap:10px;margin-top:8px;font-size:12px;color:#6b7280">
          <span id="modal-date"></span>
          <span id="modal-status"></span>
          <span>Assigned: <strong id="modal-agent"></strong></span>
        </div>
      </div>
      <button onclick="closeModal()" style="background:none;border:none;font-size:20px;cursor:pointer;color:#9ca3af;padding:0;line-height:1">&times;</button>
    </div>
    <div id="modal-msgs" style="flex:1;overflow-y:auto;padding:20px"></div>
  </div>
</div>

</body></html>"""

def cors(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return response

@app.after_request
def add_cors(response):
    return cors(response)

def fmt_dt(s):
    if not s: return ''
    try: return datetime.fromisoformat(s.replace('Z','+00:00')).strftime('%m/%d/%Y %I:%M %p')
    except: return s

def parse_ts(s):
    if not s: return None
    return datetime.fromisoformat(s.replace('Z','+00:00')).timestamp()*1000

def fetch_all_conversations(headers, from_ts, to_ts):
    conversations=[];after=None;seen=set()
    while True:
        params={'limit':50}
        if after: params['after']=after
        resp=req.get(f'{BASE}/conversations/v3/conversations/threads',headers=headers,params=params)
        resp.raise_for_status()
        data=resp.json();results=data.get('results',[])
        if not results: break
        for conv in results:
            created=conv.get('createdAt')
            if created:
                ts=parse_ts(created)
                if from_ts and ts<from_ts: continue
                if to_ts and ts>to_ts: continue
            conversations.append(conv)
        next_after=data.get('paging',{}).get('next',{}).get('after')
        if not next_after or next_after in seen: break
        seen.add(next_after);after=next_after
    return conversations

def fetch_contact(headers, contact_id):
    if not contact_id: return '','',''
    try:
        resp=req.get(f'{BASE}/crm/v3/objects/contacts/{contact_id}',headers=headers,params={'properties':'firstname,lastname,email,phone'})
        if resp.ok:
            props=resp.json().get('properties',{})
            name=f"{props.get('firstname') or ''} {props.get('lastname') or ''}".strip()
            return name,props.get('email',''),props.get('phone','')
    except: pass
    return '','',''

def resolve_agent(headers, actor_id):
    if not actor_id: return ''
    if actor_id in KNOWN_AGENTS: return KNOWN_AGENTS[actor_id]
    if actor_id in agent_cache: return agent_cache[actor_id]
    user_id=actor_id.replace('A-','') if actor_id.startswith('A-') else actor_id
    name=actor_id
    try:
        resp=req.get(f'{BASE}/crm/v3/owners/{user_id}',headers=headers)
        if resp.ok:
            d=resp.json();first=d.get('firstName','') or '';last=d.get('lastName','') or '';email=d.get('email','') or ''
            full=f'{first} {last}'.strip();name=full if full else email if email else actor_id
    except: pass
    agent_cache[actor_id]=name;return name

def fetch_messages(headers, thread_id):
    try:
        resp=req.get(f'{BASE}/conversations/v3/conversations/threads/{thread_id}/messages',headers=headers,params={'limit':100})
        resp.raise_for_status();return resp.json().get('results',[])
    except: return []

TEST_KEYWORDS = ['test','dont respond','do not respond','testing','test chat','please don\'t respond','do not interact','for roxana','dont interact']

EXCLUDED_VISITORS = [
    'test test', 'רז קלינגהופר', 'roxana tunc', 'academics@pwa.edu',
    'maricuzalonso@yahoo.com', 'maricuz', 'raz klinghoffer', 'raz klingh'
]

def is_test_conversation(conv, messages, visitor_name='', visitor_email=''):
    all_text = ' '.join((m.get('text') or m.get('body') or '').lower() for m in messages)
    if any(k in all_text for k in TEST_KEYWORDS):
        return True
    combined = ((visitor_name or '') + ' ' + (visitor_email or '')).lower()
    if any(e in combined for e in EXCLUDED_VISITORS):
        return True
    return False

def build_rows(conversations, headers):
    result=[]
    for conv in conversations:
        tid=conv.get('id','')
        contact_id=conv.get('contactId') or conv.get('associatedContactId')
        contact_name,contact_email,contact_phone=fetch_contact(headers,contact_id)
        assigned_raw=conv.get('assignedTo') or conv.get('assignedActorId') or ''
        assigned=resolve_agent(headers,assigned_raw) if assigned_raw else ''
        messages=fetch_messages(headers,tid)
        if is_test_conversation(conv, messages, contact_name, contact_email): continue
        for msg in messages:
            sender_type=msg.get('senderType') or msg.get('type') or ''
            if sender_type=='WELCOME_MESSAGE': continue
            msg_text=(msg.get('text') or msg.get('body') or msg.get('richText') or '').replace('\n',' ').strip()
            if not msg_text: continue
            sender_raw=msg.get('sender',{})
            if isinstance(sender_raw,dict):
                sender_actor=sender_raw.get('actorId','') or ''
                sender=resolve_agent(headers,sender_actor) if sender_actor.startswith('A-') else sender_type
            else: sender=sender_type
            result.append({'thread_id':tid,'created_at':conv.get('createdAt',''),'status':(conv.get('status') or '').replace('_',' ').title(),
                'visitor_name':contact_name,'visitor_email':contact_email,'visitor_phone':contact_phone,
                'assigned_to':assigned,'sender':sender,'message_time':msg.get('createdAt',''),'message':msg_text})
    return result

@app.route('/')
def index():
    if not session.get('authenticated'): return redirect('/login')
    return render_template_string(DASHBOARD_HTML)

@app.route('/login',methods=['GET','POST'])
def login():
    if request.method=='POST':
        if request.form.get('password','')==APP_PASSWORD:
            session['authenticated']=True;return redirect('/')
        return render_template_string(LOGIN_HTML,error='Incorrect password.')
    return render_template_string(LOGIN_HTML,error=None)

@app.route('/logout')
def logout():
    session.clear();return redirect('/login')

@app.route('/api/data',methods=['GET','OPTIONS'])
def api_data():
    if request.method=='OPTIONS': return cors(Response())
    if not session.get('authenticated'): return jsonify({'error':'Not authenticated'}),401
    if not HUBSPOT_TOKEN: return jsonify({'error':'HUBSPOT_TOKEN not set'}),500
    headers={'Authorization':f'Bearer {HUBSPOT_TOKEN}','Content-Type':'application/json'}
    from_date=request.args.get('from_date');to_date=request.args.get('to_date')
    from_ts=parse_ts(from_date+'T00:00:00Z') if from_date else None
    to_ts=parse_ts(to_date+'T23:59:59Z') if to_date else None
    try:
        conversations=fetch_all_conversations(headers,from_ts,to_ts)
        return jsonify(build_rows(conversations,headers))
    except Exception as e: return jsonify({'error':str(e)}),400

@app.route('/ai-analysis',methods=['POST'])
def ai_analysis():
    if not session.get('authenticated'): return jsonify({'error':'Not authenticated'}),401
    body=request.get_json()
    conversations=body.get('conversations',[])
    prompt=f"""You are analyzing live chat conversations for Pacific West Academy, a security training vocational school in California offering CESS and CST programs.

Here are {len(conversations)} recent conversations:
{json.dumps(conversations,indent=1)}

Provide a concise executive summary with:
1. Top 3 most common visitor questions or concerns
2. Any recurring issues or friction points in the chat experience
3. Notable observations about response quality or agent performance
4. 2-3 specific actionable recommendations for leadership

Keep each point to 1-2 sentences. Be specific and data-driven."""
    try:
        resp=req.post('https://api.anthropic.com/v1/messages',
            headers={'Content-Type':'application/json','x-api-key':os.environ.get('ANTHROPIC_API_KEY',''),'anthropic-version':'2023-06-01'},
            json={'model':'claude-sonnet-4-20250514','max_tokens':1000,'messages':[{'role':'user','content':prompt}]})
        data=resp.json()
        text=data.get('content',[{}])[0].get('text','No analysis returned.')
        return jsonify({'analysis':text})
    except Exception as e: return jsonify({'error':str(e)}),500

@app.route('/export',methods=['POST'])
def export():
    if not session.get('authenticated'): return jsonify({'error':'Not authenticated'}),401
    if not HUBSPOT_TOKEN: return jsonify({'error':'HUBSPOT_TOKEN not set'}),500
    body=request.get_json();from_date=body.get('from_date');to_date=body.get('to_date')
    headers={'Authorization':f'Bearer {HUBSPOT_TOKEN}','Content-Type':'application/json'}
    from_ts=parse_ts(from_date+'T00:00:00Z') if from_date else None
    to_ts=parse_ts(to_date+'T23:59:59Z') if to_date else None
    try: conversations=fetch_all_conversations(headers,from_ts,to_ts)
    except Exception as e: return jsonify({'error':str(e)}),400
    rows=build_rows(conversations,headers)
    wb=Workbook();ws=wb.active;ws.title='Live Chat Conversations'
    hf=PatternFill('solid',fgColor='1F3864');hfont=Font(color='FFFFFF',bold=True,size=11)
    cols=['Thread ID','Date','Status','Visitor Name','Visitor Email','Visitor Phone','Assigned To','Sender','Message Time','Message']
    widths=[15,20,12,22,32,18,22,22,20,80]
    for i,(h,w) in enumerate(zip(cols,widths),1):
        cell=ws.cell(row=1,column=i,value=h);cell.font=hfont;cell.fill=hf
        cell.alignment=Alignment(horizontal='center',vertical='center');ws.column_dimensions[cell.column_letter].width=w
    ws.row_dimensions[1].height=30
    fa=PatternFill('solid',fgColor='EEF2F7');fb=PatternFill('solid',fgColor='FFFFFF')
    key_map={'Thread ID':'thread_id','Date':'created_at','Status':'status','Visitor Name':'visitor_name',
        'Visitor Email':'visitor_email','Visitor Phone':'visitor_phone','Assigned To':'assigned_to',
        'Sender':'sender','Message Time':'message_time','Message':'message'}
    for i,row in enumerate(rows):
        fill=fa if i%2==0 else fb
        for j,col in enumerate(cols,1):
            cell=ws.cell(row=i+2,column=j,value=row.get(key_map[col],''))
            cell.fill=fill;cell.alignment=Alignment(wrap_text=True,vertical='top')
    ws.freeze_panes='A2';output=io.BytesIO();wb.save(output);output.seek(0)
    return send_file(output,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,download_name=f"hubspot_live_chats_{datetime.now().strftime('%Y%m%d')}.xlsx")

if __name__=='__main__':
    port=int(os.environ.get('PORT',5000));app.run(host='0.0.0.0',port=port)
